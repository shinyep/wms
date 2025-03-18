from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Q, Sum, F
from django.db import transaction
import logging
import pandas as pd
from django.http import HttpResponse, StreamingHttpResponse, FileResponse
import io
from datetime import datetime, timedelta, date
from rest_framework.renderers import BaseRenderer
from rest_framework.permissions import IsAuthenticated
from django.utils.timezone import now
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
import uuid
from django.http import Http404
from django.conf import settings
from django.db import models
from django.db.models.functions import TruncMonth, Concat
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from .models import Warehouse, WarehouseArea, WarehouseLocation, Report
from apps.inventory.models import Transaction, Inventory
from apps.product.models import Product, Unit
from .serializers import WarehouseSerializer, WarehouseAreaSerializer, WarehouseLocationSerializer, ReportSerializer
from ..user.views import WarehouseViewPermission, BasePermission
import json

logger = logging.getLogger(__name__)

class ExcelRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None
    render_style = 'binary'

    def render(self, data, media_type=None, renderer_context=None):
        return data

class WarehouseViewSet(viewsets.ModelViewSet):
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    permission_classes = [WarehouseViewPermission]
    filterset_fields = ['code', 'name', 'is_active']
    search_fields = ['code', 'name', 'address', 'contact_person']
    ordering_fields = ['id', 'code', 'name', 'created_time']
    ordering = ['-created_time']

    # 默认模板配置
    DEFAULT_TRANSACTION_COLUMNS = [
        '序号', '日期', '品项', '规格/型号', '单位', '数量', '单价', '金额', '经手人'
    ]
    
    DEFAULT_INVENTORY_COLUMNS = [
        '序号', '位置', '品项', '规格/型号', '单位', '期初库存', '累计入库', '累计出库', '库存', '单价', '库存金额'
    ]
    
    DEFAULT_SHEET_NAMES = {
        'inbound': '月入库详情',
        'outbound': '月出库详情',
        'inventory': '库存详细列表'
    }

    def retrieve(self, request, *args, **kwargs):
        """
        重写retrieve方法，增强错误处理
        """
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"获取仓库信息失败: {str(e)}", exc_info=True)
            return Response({"error": f"获取仓库信息失败: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_object(self):
        """
        重写get_object方法，确保能正确获取仓库对象
        """
        try:
            pk = self.kwargs.get('pk')
            if not pk:
                raise Http404("未提供仓库ID")
                
            warehouse = Warehouse.objects.get(pk=pk)
            self.check_object_permissions(self.request, warehouse)
            return warehouse
        except Warehouse.DoesNotExist:
            raise Http404("仓库不存在")
        except Exception as e:
            logger.error(f"获取仓库对象失败: {str(e)}", exc_info=True)
            raise Http404(f"获取仓库信息失败: {str(e)}")

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """获取仪表盘数据"""
        try:
            logger.info("开始获取仪表盘数据")
            
            # 获取基础统计数据
            warehouse_count = Warehouse.objects.filter(is_active=True).count()
            logger.info(f"Active warehouse count: {warehouse_count}")

            area_count = WarehouseArea.objects.filter(is_active=True).count()
            logger.info(f"Active area count: {area_count}")

            location_count = WarehouseLocation.objects.filter(is_active=True).count()
            used_location_count = WarehouseLocation.objects.filter(is_active=True, is_empty=False).count()
            try:
                location_usage = round(used_location_count / location_count * 100, 2) if location_count > 0 else 0
            except Exception as e:
                logger.error(f"计算库位使用率时出错: {str(e)}")
                location_usage = 0
            logger.info(f"Location stats - total: {location_count}, used: {used_location_count}, usage: {location_usage}%")

            # 获取每个仓库的库位统计
            warehouse_stats = []
            try:
                warehouses = Warehouse.objects.filter(is_active=True)
                for warehouse in warehouses:
                    try:
                        location_count = WarehouseLocation.objects.filter(
                            area__warehouse=warehouse,
                            is_active=True
                        ).count()
                        warehouse_stats.append({
                            'name': warehouse.name,
                            'location_count': location_count
                        })
                    except Exception as e:
                        logger.error(f"获取仓库 {warehouse.name} 的库位统计时出错: {str(e)}")
                        continue
                logger.info(f"Warehouse stats: {warehouse_stats}")
            except Exception as e:
                logger.error(f"获取仓库统计时出错: {str(e)}")

            # 获取库区类型分布
            area_stats = []
            area_types = WarehouseArea.objects.filter(is_active=True).values('area_type').annotate(
                count=Count('id')
            )
            for area_type in area_types:
                area_stats.append({
                    'name': area_type['area_type'] or '未分类',
                    'count': area_type['count']
                })
            logger.info(f"Area stats: {area_stats}")

            response_data = {
                'warehouse_count': warehouse_count,
                'area_count': area_count,
                'location_count': location_count,
                'location_usage': location_usage,
                'warehouse_stats': warehouse_stats,
                'area_stats': area_stats
            }
            logger.info(f"Dashboard data retrieved successfully: {response_data}")
            return Response(response_data)

        except Exception as e:
            logger.error(f"Error retrieving dashboard data: {str(e)}", exc_info=True)
            error_message = f"获取仪表盘数据失败: {str(e)}"
            logger.error(error_message)
            return Response({
                'warehouse_count': 0,
                'area_count': 0,
                'location_count': 0,
                'location_usage': 0,
                'warehouse_stats': [],
                'area_stats': [],
                'error': error_message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def monthly_report(self, request):
        """
        获取指定仓库的月度报表
        支持Excel和JSON格式的响应
        可通过format参数指定返回格式：excel或json
        增加save参数，用于保存报表数据
        增加check_only参数，仅检查是否有数据，不返回详细数据
        """
        try:
            # 获取请求参数
            warehouse_id = request.query_params.get('warehouse_id')
            month = request.query_params.get('month')  # 格式: YYYY-MM
            response_format = request.query_params.get('format', 'excel')  # 默认为excel
            save_report = request.query_params.get('save', 'false').lower() == 'true'  # 是否保存报表
            limit = int(request.query_params.get('limit', 5000))  # 数据限制，默认5000条
            check_only = request.query_params.get('check_only', 'false').lower() == 'true'  # 是否仅检查
            
            # 添加日志
            logger.info(f"开始处理月度报表请求: warehouse_id={warehouse_id}, month={month}, format={response_format}, save={save_report}, limit={limit}, check_only={check_only}")

            # 检查必填参数
            if not warehouse_id or not month:
                return Response({"error": "warehouse_id和month是必填参数"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({"error": "仓库不存在"}, status=status.HTTP_404_NOT_FOUND)
            except ValueError:
                return Response({"error": "仓库ID格式错误"}, status=status.HTTP_400_BAD_REQUEST)

            # 解析月份并计算日期范围
            try:
                year, month = month.split('-')
                year, month = int(year), int(month)

                # 验证月份范围
                if not (1 <= month <= 12):
                    return Response({"error": "月份必须在1-12之间"}, status=status.HTTP_400_BAD_REQUEST)

                report_date = f"{year}-{month:02d}-01"
                logger.info(f"生成报表日期: {report_date}")
                
                # 先查找已存在的报表
                report = Report.objects.filter(
                    warehouse=warehouse,
                    report_date=report_date
                ).first()
                
                # 如果需要JSON格式且报表已存在
                if response_format.lower() == 'json':
                    logger.info("准备返回JSON格式数据")
                    
                    if report:
                        logger.info(f"找到已存在的报表: id={report.id}, data_keys={report.data.keys() if report.data else None}")
                        # 只检查数据是否存在
                        if check_only:
                            return Response({
                                "has_data": True,
                                "report_id": report.id
                            })
                        
                        # 直接返回报表数据
                        # 确保含有所有必要的字段
                        report_data = report.data.copy() if report.data else {}
                        if 'inbound' not in report_data:
                            report_data['inbound'] = []
                        if 'outbound' not in report_data:
                            report_data['outbound'] = []
                        if 'inventory' not in report_data:
                            report_data['inventory'] = []
                            
                        logger.info(f"返回报表数据: inbound={len(report_data['inbound'])}条, outbound={len(report_data['outbound'])}条, inventory={len(report_data['inventory'])}条")
                        return Response(report_data)
                    else:
                        logger.info("未找到报表，返回空数据")
                        # 返回空数据结构
                        empty_data = {
                            'inbound': [],
                            'outbound': [],
                            'inventory': []
                        }
                        return Response(empty_data)
                
                # 对于Excel格式，生成Excel文件
                elif response_format.lower() == 'excel':
                    logger.info("准备生成Excel报表")
                    
                    if report and report.data:
                        logger.info(f"使用已存在的报表生成Excel: id={report.id}")
                        return self.generate_excel_report(report, warehouse, year, month)
                    else:
                        logger.info("生成空的Excel模板")
                        return self.generate_empty_excel_template(warehouse, year, month)
                else:
                    return Response({"error": "不支持的格式，请使用json或excel"}, status=status.HTTP_400_BAD_REQUEST)
                    
            except Exception as e:
                logger.error(f"处理月度报表数据时发生错误: {str(e)}", exc_info=True)
                return Response({"error": f"处理月度报表数据时发生错误: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        except Exception as e:
            logger.error(f"处理月度报表请求时发生错误: {str(e)}", exc_info=True)
            return Response({"error": f"处理月度报表请求时发生错误: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def generate_empty_excel_template(self, warehouse, year, month):
        """
        生成空的Excel模板
        """
        try:
            excel_file = io.BytesIO()
            
            # 创建工作簿
            workbook = pd.ExcelWriter(excel_file, engine='openpyxl')
            
            # 创建入库明细表
            inbound_df = pd.DataFrame(columns=self.DEFAULT_TRANSACTION_COLUMNS)
            inbound_df.to_excel(workbook, sheet_name=self.DEFAULT_SHEET_NAMES['inbound'], index=False)
            
            # 创建出库明细表
            outbound_df = pd.DataFrame(columns=self.DEFAULT_TRANSACTION_COLUMNS)
            outbound_df.to_excel(workbook, sheet_name=self.DEFAULT_SHEET_NAMES['outbound'], index=False)
            
            # 创建库存明细表
            inventory_df = pd.DataFrame(columns=self.DEFAULT_INVENTORY_COLUMNS)
            inventory_df.to_excel(workbook, sheet_name=self.DEFAULT_SHEET_NAMES['inventory'], index=False)
            
            # 设置每个sheet的格式
            for sheet_name in self.DEFAULT_SHEET_NAMES.values():
                worksheet = workbook.sheets[sheet_name]
                # 设置标题行格式
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Workaround for Python context behavior
            workbook._cleanup()
            workbook.close()
            
            excel_file.seek(0)
            
            # 生成文件名
            filename = f"{warehouse.name}_月度报表_{year}年{month:02d}月.xlsx"
            
            # 返回文件
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"生成Excel模板失败: {str(e)}", exc_info=True)
            raise e
            
    def generate_excel_report(self, report, warehouse, year, month):
        """
        根据报表数据生成Excel文件
        """
        try:
            excel_file = io.BytesIO()
            
            # 创建工作簿
            workbook = pd.ExcelWriter(excel_file, engine='openpyxl')
            
            # 处理入库明细数据
            inbound_data = report.data.get('inbound', [])
            if inbound_data:
                # 移除id字段，不显示在Excel中
                for item in inbound_data:
                    if 'id' in item:
                        del item['id']
                inbound_df = pd.DataFrame(inbound_data)
            else:
                inbound_df = pd.DataFrame(columns=self.DEFAULT_TRANSACTION_COLUMNS)
            
            inbound_df.to_excel(workbook, sheet_name=self.DEFAULT_SHEET_NAMES['inbound'], index=False)
            
            # 处理出库明细数据
            outbound_data = report.data.get('outbound', [])
            if outbound_data:
                # 移除id字段，不显示在Excel中
                for item in outbound_data:
                    if 'id' in item:
                        del item['id']
                outbound_df = pd.DataFrame(outbound_data)
            else:
                outbound_df = pd.DataFrame(columns=self.DEFAULT_TRANSACTION_COLUMNS)
            
            outbound_df.to_excel(workbook, sheet_name=self.DEFAULT_SHEET_NAMES['outbound'], index=False)
            
            # 处理库存明细数据
            inventory_data = report.data.get('inventory', [])
            if inventory_data:
                # 移除id字段，不显示在Excel中
                for item in inventory_data:
                    if 'id' in item:
                        del item['id']
                inventory_df = pd.DataFrame(inventory_data)
            else:
                inventory_df = pd.DataFrame(columns=self.DEFAULT_INVENTORY_COLUMNS)
            
            inventory_df.to_excel(workbook, sheet_name=self.DEFAULT_SHEET_NAMES['inventory'], index=False)
            
            # 设置每个sheet的格式
            for sheet_name in self.DEFAULT_SHEET_NAMES.values():
                worksheet = workbook.sheets[sheet_name]
                # 设置标题行格式
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Workaround for Python context behavior
            workbook._cleanup()
            workbook.close()
            
            excel_file.seek(0)
            
            # 生成文件名
            filename = f"{warehouse.name}_月度报表_{year}年{month:02d}月.xlsx"
            
            # 返回文件
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"生成Excel报表失败: {str(e)}", exc_info=True)
            raise e

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """导入Excel文件"""
        try:
            excel_file = request.FILES.get('file')
            if not excel_file:
                return Response({'error': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 从文件名获取仓库名称（去掉.xlsx扩展名）
            warehouse_name = excel_file.name
            if warehouse_name.lower().endswith('.xlsx'):
                warehouse_name = warehouse_name[:-5]
            
            # 生成仓库编码（使用时间戳）
            warehouse_code = f"WH{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # 创建新仓库
            try:
                warehouse = Warehouse.objects.create(
                    name=warehouse_name,
                    code=warehouse_code,
                    is_active=True
                )
            except Exception as e:
                return Response({'error': f'创建仓库失败: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 读取所有sheet
            excel_data = pd.read_excel(excel_file, sheet_name=None)
            
            # 验证sheet名称是否符合默认模板
            required_sheets = set(self.DEFAULT_SHEET_NAMES.values())
            if not all(sheet in excel_data for sheet in required_sheets):
                missing_sheets = required_sheets - set(excel_data.keys())
                return Response(
                    {'error': f'Excel文件缺少必要的sheet: {", ".join(missing_sheets)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            success_count = {
                'outbound': 0,
                'inbound': 0,
                'inventory': 0
            }
            error_records = []
            
            # 获取当前用户作为操作人
            operator = request.user
            
            # 处理月出库详情
            outbound_df = excel_data[self.DEFAULT_SHEET_NAMES['outbound']]
            if not all(col in outbound_df.columns for col in self.DEFAULT_TRANSACTION_COLUMNS):
                missing_cols = set(self.DEFAULT_TRANSACTION_COLUMNS) - set(outbound_df.columns)
                error_records.append(f'出库明细缺少必要的列: {", ".join(missing_cols)}')
            else:
                for index, row in outbound_df.iterrows():
                    try:
                        # 处理出库数据
                        product_name = str(row['品项']).strip()
                        spec = str(row['规格/型号']).strip()
                        unit_name = str(row['单位']).strip()
                        
                        # 获取或创建商品
                        try:
                            product = Product.objects.get(name=product_name, spec=spec)
                        except Product.DoesNotExist:
                            # 获取或创建单位
                            try:
                                unit = Unit.objects.get(name=unit_name)
                            except Unit.DoesNotExist:
                                unit = Unit.objects.create(
                                    name=unit_name,
                                    code=f"U{datetime.now().strftime('%Y%m%d%H%M%S')}"
                                )
                            
                            # 创建商品
                            product = Product.objects.create(
                                name=product_name,
                                spec=spec,
                                unit=unit,
                                code=f"P{datetime.now().strftime('%Y%m%d%H%M%S')}",
                                is_active=True
                            )
                        
                        # 创建出库记录
                        transaction_data = {
                            'warehouse': warehouse,
                            'product': product,
                            'transaction_type': 'OUT',
                            'quantity': float(row['数量']),
                            'unit_price': float(row['单价']),
                            'amount': float(row['金额']),
                            'operator': operator,
                            'transaction_date': pd.to_datetime(row['日期']),
                            'status': 'completed',
                            'is_active': True
                        }
                        Transaction.objects.create(**transaction_data)
                        success_count['outbound'] += 1
                    except Exception as e:
                        error_records.append(f"出库明细第 {index + 2} 行处理失败: {str(e)}")
            
            # 处理月入库详情
            inbound_df = excel_data[self.DEFAULT_SHEET_NAMES['inbound']]
            if not all(col in inbound_df.columns for col in self.DEFAULT_TRANSACTION_COLUMNS):
                missing_cols = set(self.DEFAULT_TRANSACTION_COLUMNS) - set(inbound_df.columns)
                error_records.append(f'入库明细缺少必要的列: {", ".join(missing_cols)}')
            else:
                for index, row in inbound_df.iterrows():
                    try:
                        # 处理入库数据
                        product_name = str(row['品项']).strip()
                        spec = str(row['规格/型号']).strip()
                        unit_name = str(row['单位']).strip()
                        
                        # 获取或创建商品
                        try:
                            product = Product.objects.get(name=product_name, spec=spec)
                        except Product.DoesNotExist:
                            # 获取或创建单位
                            try:
                                unit = Unit.objects.get(name=unit_name)
                            except Unit.DoesNotExist:
                                unit = Unit.objects.create(
                                    name=unit_name,
                                    code=f"U{datetime.now().strftime('%Y%m%d%H%M%S')}"
                                )
                            
                            # 创建商品
                            product = Product.objects.create(
                                name=product_name,
                                spec=spec,
                                unit=unit,
                                code=f"P{datetime.now().strftime('%Y%m%d%H%M%S')}",
                                is_active=True
                            )
                        
                        # 创建入库记录
                        transaction_data = {
                            'warehouse': warehouse,
                            'product': product,
                            'transaction_type': 'IN',
                            'quantity': float(row['数量']),
                            'unit_price': float(row['单价']),
                            'amount': float(row['金额']),
                            'operator': operator,
                            'transaction_date': pd.to_datetime(row['日期']),
                            'status': 'completed',
                            'is_active': True
                        }
                        Transaction.objects.create(**transaction_data)
                        success_count['inbound'] += 1
                    except Exception as e:
                        error_records.append(f"入库明细第 {index + 2} 行处理失败: {str(e)}")
            
            # 处理库存明细
            inventory_df = excel_data[self.DEFAULT_SHEET_NAMES['inventory']]
            if not all(col in inventory_df.columns for col in self.DEFAULT_INVENTORY_COLUMNS):
                missing_cols = set(self.DEFAULT_INVENTORY_COLUMNS) - set(inventory_df.columns)
                error_records.append(f'库存明细缺少必要的列: {", ".join(missing_cols)}')
            else:
                for index, row in inventory_df.iterrows():
                    try:
                        # 处理库存数据
                        product_name = str(row['品项']).strip()
                        spec = str(row['规格/型号']).strip()
                        unit_name = str(row['单位']).strip()
                        location_code = str(row['位置']).strip()
                        
                        # 获取或创建商品
                        try:
                            product = Product.objects.get(name=product_name, spec=spec)
                        except Product.DoesNotExist:
                            # 获取或创建单位
                            try:
                                unit = Unit.objects.get(name=unit_name)
                            except Unit.DoesNotExist:
                                unit = Unit.objects.create(
                                    name=unit_name,
                                    code=f"U{datetime.now().strftime('%Y%m%d%H%M%S')}"
                                )
                            
                            # 创建商品
                            product = Product.objects.create(
                                name=product_name,
                                spec=spec,
                                unit=unit,
                                code=f"P{datetime.now().strftime('%Y%m%d%H%M%S')}",
                                is_active=True
                            )
                        
                        # 获取或创建库位
                        try:
                            location = WarehouseLocation.objects.get(code=location_code, area__warehouse=warehouse)
                        except WarehouseLocation.DoesNotExist:
                            # 创建默认库区
                            default_area = WarehouseArea.objects.create(
                                warehouse=warehouse,
                                name='默认库区',
                                code=f"A{datetime.now().strftime('%Y%m%d%H%M%S')}",
                                is_active=True
                            )
                            # 创建库位
                            location = WarehouseLocation.objects.create(
                                area=default_area,
                                code=location_code,
                                name=f"库位{location_code}",
                                is_active=True
                            )
                        
                        # 创建库存记录
                        inventory_data = {
                            'warehouse': warehouse,
                            'product': product,
                            'location': location,
                            'quantity': float(row['库存']),
                            'unit_price': float(row['单价']),
                            'amount': float(row['库存金额']),
                            'status': 'normal',
                            'is_active': True
                        }
                        Inventory.objects.create(**inventory_data)
                        success_count['inventory'] += 1
                    except Exception as e:
                        error_records.append(f"库存明细第 {index + 2} 行处理失败: {str(e)}")
            
            # 返回导入结果
            response_data = {
                'message': f'Excel导入完成，成功创建仓库 "{warehouse_name}"',
                'warehouse_id': warehouse.id,
                'warehouse_name': warehouse.name,
                'warehouse_code': warehouse.code,
                'success_count': {
                    '出库记录': success_count['outbound'],
                    '入库记录': success_count['inbound'],
                    '库存记录': success_count['inventory']
                }
            }
            if error_records:
                response_data['errors'] = error_records
            
            return Response(response_data)
            
        except Exception as e:
            return Response(
                {'error': f'Excel导入失败: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['delete'])
    def delete_record(self, request):
        """删除报表记录，同时清理关联的入库和出库记录"""
        try:
            print("===== 收到删除记录请求 =====")
            print(f"查询参数: {request.query_params}")
            
            warehouse_id = request.query_params.get('warehouse_id')
            month = request.query_params.get('month')
            record_type = request.query_params.get('record_type')
            record_id = request.query_params.get('record_id')

            if not all([warehouse_id, month, record_type, record_id]):
                print(f"缺少必要参数: warehouse_id={warehouse_id}, month={month}, record_type={record_type}, record_id={record_id}")
                return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                with transaction.atomic():  # 使用事务确保数据一致性
                    warehouse = Warehouse.objects.get(id=warehouse_id)
                    year, month_num = month.split('-')
                    report_date = f"{year}-{month_num}-01"
                    
                    report = Report.objects.get(
                        warehouse=warehouse,
                        report_date=report_date
                    )
                    
                    # 记录删除前的数据量
                    before_delete = {
                        'inbound': len(report.data.get('inbound', [])),
                        'outbound': len(report.data.get('outbound', [])),
                        'inventory': len(report.data.get('inventory', []))
                    }
                    
                    records = report.data.get(record_type, [])
                    if not records:
                        return Response({'error': f'没有找到{record_type}类型的记录'}, status=status.HTTP_404_NOT_FOUND)
                    
                    # 初始化删除的相关记录计数
                    related_deleted = {'inbound': 0, 'outbound': 0, 'transactions': 0}
                    
                    # 查找目标记录
                    target_item = None
                    target_index = -1
                    for i, item in enumerate(records):
                        if item.get('id') == record_id:
                            target_item = item
                            target_index = i
                            break
                    
                    if target_item is None:
                        print(f"未找到记录ID: {record_id}")
                        return Response({'error': f'未找到记录(ID: {record_id})'}, status=status.HTTP_404_NOT_FOUND)
                    
                    # 获取品项和规格
                    item_name = target_item.get('品项', '')
                    item_spec = target_item.get('规格/型号', '')
                    print(f"要删除的记录: 类型={record_type}, 品项={item_name}, 规格={item_spec}")
                    
                    # 查找相关的出入库记录
                    if record_type == 'inventory':
                        # 处理库存明细记录删除
                        # 1. 找到并删除相关入库记录
                        inbound_records = report.data.get('inbound', [])
                        original_inbound_count = len(inbound_records)
                        filtered_inbound = []
                        
                        for r in inbound_records:
                            # 精确匹配品项和规格
                            if r.get('品项') == item_name and r.get('规格/型号') == item_spec:
                                print(f"删除相关入库记录: {r.get('id', 'no-id')}, 日期={r.get('日期', 'unknown')}")
                            else:
                                filtered_inbound.append(r)
                        
                        report.data['inbound'] = filtered_inbound
                        related_deleted['inbound'] = original_inbound_count - len(filtered_inbound)
                        
                        # 2. 找到并删除相关出库记录
                        outbound_records = report.data.get('outbound', [])
                        original_outbound_count = len(outbound_records)
                        filtered_outbound = []
                        
                        for r in outbound_records:
                            # 精确匹配品项和规格
                            if r.get('品项') == item_name and r.get('规格/型号') == item_spec:
                                print(f"删除相关出库记录: {r.get('id', 'no-id')}, 日期={r.get('日期', 'unknown')}")
                            else:
                                filtered_outbound.append(r)
                        
                        report.data['outbound'] = filtered_outbound
                        related_deleted['outbound'] = original_outbound_count - len(filtered_outbound)
                        
                        # 3. 删除数据库中对应的Transaction记录
                        from apps.inventory.models import Transaction, Product
                        
                        # 先查找完全匹配的产品
                        products = Product.objects.filter(name=item_name)
                        if not products.exists():
                            # 如果找不到完全匹配的，尝试模糊匹配但更加谨慎
                            products = Product.objects.filter(name__icontains=item_name)
                            if products.count() > 3:  # 避免匹配过多产品导致误删
                                print(f"模糊匹配到过多产品({products.count()}个)，不执行删除")
                                products = Product.objects.none()
                        
                        if products.exists():
                            # 构建查询条件
                            query_params = {
                                'warehouse_id': warehouse_id,
                                'product__in': products,
                                'transaction_date__year': int(year),
                                'transaction_date__month': int(month_num)
                            }
                            
                            # 精确匹配规格
                            if item_spec:
                                query_params['spec'] = item_spec
                            
                            transactions = Transaction.objects.filter(**query_params)
                            
                            # 记录将删除的事务详情
                            transaction_count = transactions.count()
                            for t in transactions:
                                print(f"将删除Transaction: ID={t.id}, 类型={t.transaction_type}, 产品={t.product.name}, 规格={t.spec}, 日期={t.transaction_date}")
                            
                            if transaction_count > 0:
                                transactions.delete()
                                related_deleted['transactions'] = transaction_count
                                print(f"成功删除{transaction_count}条Transaction记录")
                            else:
                                print(f"未找到与品项[{item_name}]规格[{item_spec}]匹配的Transaction记录")
                        else:
                            print(f"未找到与品项[{item_name}]匹配的Product记录")
                    
                    elif record_type == 'inbound' or record_type == 'outbound':
                        # 处理入库/出库记录删除
                        from apps.inventory.models import Transaction, Product
                        
                        transaction_type = 'IN' if record_type == 'inbound' else 'OUT'
                        item_date = target_item.get('日期')
                        print(f"删除{record_type}记录: 品项={item_name}, 规格={item_spec}, 日期={item_date}")
                        
                        # 查找匹配的产品
                        products = Product.objects.filter(name=item_name)
                        if not products.exists():
                            products = Product.objects.filter(name__icontains=item_name)
                            if products.count() > 3:
                                print(f"模糊匹配到过多产品({products.count()}个)，不执行删除")
                                products = Product.objects.none()
                        if products.exists():
                            # 构建基本查询条件
                            query_params = {
                                'warehouse_id': warehouse_id,
                                'product__in': products,
                                'transaction_type': transaction_type
                            }
                            
                            # 精确匹配规格
                            if item_spec:
                                query_params['spec'] = item_spec
                                
                            # 匹配日期(如果有)
                            if item_date:
                                try:
                                    transaction_date = datetime.strptime(item_date, '%Y-%m-%d').date()
                                    query_params['transaction_date'] = transaction_date
                                    print(f"匹配日期: {transaction_date}")
                                except Exception as e:
                                    print(f"日期格式转换失败: {e}, 原始日期={item_date}")
                                    # 使用年月匹配
                                    query_params['transaction_date__year'] = int(year)
                                    query_params['transaction_date__month'] = int(month_num)
                            else:
                                # 使用年月匹配
                                query_params['transaction_date__year'] = int(year)
                                query_params['transaction_date__month'] = int(month_num)
                            
                            # 匹配数量(如果有)
                            if 'quantity' in target_item and target_item['quantity']:
                                query_params['quantity'] = target_item['quantity']
                                
                            # 查找并删除匹配的Transaction记录
                            transactions = Transaction.objects.filter(**query_params)
                            
                            # 记录详情
                            transaction_count = transactions.count()
                            for t in transactions:
                                print(f"将删除Transaction: ID={t.id}, 类型={t.transaction_type}, 产品={t.product.name}, 规格={t.spec}, 日期={t.transaction_date}")
                            
                            if transaction_count > 0:
                                transactions.delete()
                                related_deleted['transactions'] = transaction_count
                                print(f"成功删除{transaction_count}条Transaction记录")
                            else:
                                print(f"未找到与{record_type}记录匹配的Transaction")
                        else:
                            print(f"未找到与品项[{item_name}]匹配的Product记录")
                    
                    # 删除报表中的目标记录
                    if target_index >= 0:
                        records.pop(target_index)
                        report.data[record_type] = records
                        print(f"从报表中删除了{record_type}记录")
                    
                    # 保存更改
                    report.save()
                    
                    # 记录删除后的数据量
                    after_delete = {
                        'inbound': len(report.data.get('inbound', [])),
                        'outbound': len(report.data.get('outbound', [])),
                        'inventory': len(report.data.get('inventory', []))
                    }
                    
                    # 检查是否有变化
                    has_changes = (
                        before_delete['inbound'] != after_delete['inbound'] or
                        before_delete['outbound'] != after_delete['outbound'] or
                        before_delete['inventory'] != after_delete['inventory'] or
                        related_deleted['transactions'] > 0
                    )
                    
                    if not has_changes:
                        print("删除操作没有产生任何变化")
                        return Response({'error': '删除操作没有产生任何变化'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    
                    # 从数据库重新读取确认变更已保存
                    report.refresh_from_db()
                    
                    # 记录日志
                    print(f"删除成功: before={before_delete}, after={after_delete}, related_deleted={related_deleted}")
                    
                    return Response({
                        'status': 'success',
                        'message': '删除成功',
                        'data': {
                            'before': before_delete,
                            'after': after_delete,
                            'related_deleted': related_deleted
                        }
                    })
                
            except Warehouse.DoesNotExist:
                return Response({'error': '仓库不存在'}, status=status.HTTP_404_NOT_FOUND)
            except Report.DoesNotExist:
                return Response({'error': '报表不存在'}, status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                import traceback
                print(f"删除记录失败: {str(e)}")
                print(traceback.format_exc())
                return Response({'error': f'删除记录失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            import traceback
            print(f"系统错误: {str(e)}")
            print(traceback.format_exc())
            return Response({'error': f'系统错误: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post', 'put'])
    def update_monthly_report(self, request):
        """
        更新月度报表数据
        """
        try:
            # 明确检查用户权限 - 确保1级账号可以操作
            user_level = request.user.level if hasattr(request, 'user') else 0
            if user_level < 1:
                return Response({'error': '权限不足，需要高级及以上权限'}, status=status.HTTP_403_FORBIDDEN)
            
            warehouse_id = request.data.get('warehouse_id')
            month = request.data.get('month')
            data = request.data.get('data', {})
            inbound_data = data.get('inbound', [])
            outbound_data = data.get('outbound', [])
            inventory_data = data.get('inventory', [])
            
            if not warehouse_id or not month:
                return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({'error': '仓库不存在'}, status=status.HTTP_404_NOT_FOUND)
            
            # 检查月份格式
            try:
                year, month_num = month.split('-')
                year = int(year)
                month_num = int(month_num)
                if not (1 <= month_num <= 12):
                    return Response({'error': '月份格式错误，应为YYYY-MM'}, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({'error': '月份格式错误，应为YYYY-MM'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 生成唯一报表标识
            report_key = f"{warehouse.id}_{year}_{month_num}"
            
            # 首先尝试通过key查找
            report = None
            try:
                report = Report.objects.filter(key=report_key).first()
            except Exception as e:
                logger.error(f"通过key查找报表失败: {str(e)}")
                
            # 如果通过key没找到，尝试其他条件查找
            if not report:
                try:
                    report_date = f"{year}-{month_num:02d}-01"
                    reports = Report.objects.filter(
                        warehouse=warehouse,
                        report_date=report_date
                    )
                    if reports.exists():
                        report = reports.first()
                        # 更新key
                        report.key = report_key
                except Exception as e:
                    logger.error(f"通过日期查找报表失败: {str(e)}")
            
            # 如果仍未找到，创建新报表
            if not report:
                report_date = f"{year}-{month_num:02d}-01"
                report = Report.objects.create(
                    title=f"{warehouse.name} {year}年{month_num}月报表",
                    key=report_key,
                    warehouse=warehouse,
                    report_date=report_date,
                    data={},
                    creator=request.user if hasattr(request, 'user') and request.user.is_authenticated else None
                )
            
            # 更新报表数据
            report_data = {}
            
            # 确保数据中的每个记录都有唯一ID
            def ensure_record_ids(records, prefix):
                result = []
                seen_ids = set()
                for record in records:
                    if not record:  # 跳过空记录
                        continue
                    # 如果记录没有ID或ID重复，生成新ID
                    if 'id' not in record or record['id'] in seen_ids:
                        record['id'] = f"{prefix}_{uuid.uuid4()}"
                    seen_ids.add(record['id'])
                    result.append(record)
                return result
            
            # 处理入库、出库和库存数据
            inbound_data = ensure_record_ids(inbound_data, 'in')
            outbound_data = ensure_record_ids(outbound_data, 'out')
            inventory_data = ensure_record_ids(inventory_data, 'inv')
            
            # 记录更新前的数据量
            before_update = {
                'inbound': len(report.data.get('inbound', [])) if report.data else 0,
                'outbound': len(report.data.get('outbound', [])) if report.data else 0,
                'inventory': len(report.data.get('inventory', [])) if report.data else 0
            }
            
            # 完全替换入库、出库和库存数据
            report_data['inbound'] = inbound_data
            report_data['outbound'] = outbound_data
            report_data['inventory'] = inventory_data
            
            # 记录更新后的数据量
            after_update = {
                'inbound': len(inbound_data),
                'outbound': len(outbound_data),
                'inventory': len(inventory_data)
            }
            
            # 检查数据变化
            logger.info(f"数据更新前: {before_update}")
            logger.info(f"数据更新后: {after_update}")
            
            # 保存更新后的数据
            report.data = report_data
            report.save()
            
            # 强制从数据库重新读取数据以验证保存是否成功
            report.refresh_from_db()
            saved_data = report.data
            
            # 验证保存的数据
            if (len(saved_data.get('inventory', [])) != len(inventory_data) or
                len(saved_data.get('inbound', [])) != len(inbound_data) or
                len(saved_data.get('outbound', [])) != len(outbound_data)):
                logger.error("数据保存验证失败")
                return Response({
                    'error': '数据保存验证失败，请重试'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # 记录操作日志
            logger.info(f"更新月度报表成功: warehouse={warehouse.name}, month={month}, "
                       f"inventory_count={len(inventory_data)}, "
                       f"inbound_count={len(inbound_data)}, "
                       f"outbound_count={len(outbound_data)}")
            
            return Response({
                'status': 'success', 
                'message': '报表数据更新成功',
                'data': {
                    'inbound': inbound_data,
                    'outbound': outbound_data,
                    'inventory': inventory_data
                }
            }, status=status.HTTP_200_OK)
        except Exception as e:
            import traceback
            logger.error(f"更新月度报表出错: {str(e)}\n{traceback.format_exc()}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def create_auto_monthly_report(self, request):
        """
        自动创建新月度报表，基于上个月的报表数据
        
        可用于两种场景：
        1. 每月1日自动创建当月报表
        2. 手动补创建当月或指定月份的报表
        
        处理逻辑：
        - 获取上个月的库存数据作为当月的期初库存
        - 入库和出库记录初始为空
        - 如果已存在指定月份的报表，则更新而不是创建新报表
        """
        try:
            warehouse_id = request.data.get('warehouse_id')
            target_month = request.data.get('month')  # 目标月份，格式: YYYY-MM
            
            if not warehouse_id or not target_month:
                return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({'error': '仓库不存在'}, status=status.HTTP_404_NOT_FOUND)
            
            # 检查月份格式并获取上个月的日期
            try:
                target_date = datetime.strptime(f"{target_month}-01", "%Y-%m-%d")
                # 计算上个月的日期
                first_day_of_month = target_date.replace(day=1)
                last_day_of_prev_month = first_day_of_month - timedelta(days=1)
                prev_month = last_day_of_prev_month.strftime("%Y-%m")
            except ValueError:
                return Response({'error': '月份格式错误，应为YYYY-MM'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 获取上个月的报表数据
            try:
                # 查询方式类似于monthly_report方法
                year, month_num = prev_month.split('-')
                year = int(year)
                month_num = int(month_num)
                
                # 获取上月报表数据
                prev_month_data = self.get_monthly_report_data(warehouse, year, month_num)
                
                if not prev_month_data.get('inventory'):
                    return Response({
                        'error': f'上个月({prev_month})没有库存数据，无法自动创建本月报表'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # 创建本月报表：入库和出库为空列表，库存使用上月期末库存作为期初
                new_month_data = {
                    'inbound': [],  # 新月份初始没有入库记录
                    'outbound': [], # 新月份初始没有出库记录
                    'inventory': []
                }
                
                # 复制上月库存数据作为本月的期初库存
                for item in prev_month_data.get('inventory', []):
                    new_inventory_item = item.copy()
                    # 重置统计数据
                    new_inventory_item['期初库存'] = item['库存']  # 上月期末库存作为本月期初
                    new_inventory_item['累计入库'] = 0
                    new_inventory_item['累计出库'] = 0
                    new_inventory_item['库存'] = item['库存']  # 初始库存等于期初库存
                    new_month_data['inventory'].append(new_inventory_item)
                
                # 保存新报表
                report_title = f"{warehouse.name}_{target_month}月度报表"
                report_date = datetime.strptime(f"{target_month}-01", "%Y-%m-%d").date()
                
                # 检查是否已经存在此月报表
                existing_report = Report.objects.filter(
                    warehouse=warehouse,
                    report_date__year=target_date.year,
                    report_date__month=target_date.month
                ).first()
                
                if existing_report:
                    # 更新已有报表
                    existing_report.data = new_month_data
                    existing_report.updated_at = datetime.now()
                    existing_report.save()
                    message = f'已更新{target_month}月度报表'
                else:
                    # 创建新报表
                    Report.objects.create(
                        title=report_title,
                        warehouse=warehouse,
                        report_date=report_date,
                        data=new_month_data,
                        creator=request.user if request.user.is_authenticated else None,
                        description=f"系统自动创建的{target_month}月度报表，基于{prev_month}月数据"
                    )
                    message = f'已自动创建{target_month}月度报表'
                
                return Response({
                    'status': 'success',
                    'message': message,
                    'data': {
                        'warehouse_id': warehouse.id,
                        'month': target_month,
                        'prev_month': prev_month,
                        'inventory_count': len(new_month_data['inventory'])
                    }
                }, status=status.HTTP_200_OK)
                
            except Exception as e:
                return Response({
                    'error': f'获取上月报表数据失败: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update_initial_inventory_for_month(self, warehouse_id, month):
        """
        更新指定月份报表的期初库存数据，基于上月的期末库存
        
        参数:
            warehouse_id: 仓库ID
            month: 目标月份，格式：YYYY-MM
        """
        try:
            # 参数校验
            if not warehouse_id or not month:
                return {'error': '缺少必要参数'}
                
            # 解析月份
            try:
                year, month_num = month.split('-')
                year = int(year)
                month_num = int(month_num)
                
                # 计算上个月的日期
                first_day_of_month = datetime(year, month_num, 1)
                last_day_of_prev_month = first_day_of_month - timedelta(days=1)
                prev_month = last_day_of_prev_month.strftime("%Y-%m")
                prev_year = last_day_of_prev_month.year
                prev_month_num = last_day_of_prev_month.month
                
                print(f"正在处理月份: 当前月={year}-{month_num}, 上个月={prev_year}-{prev_month_num}")
            except (ValueError, IndexError) as e:
                print(f"月份格式解析错误: {str(e)}")
                return {'error': f'月份格式错误，应为YYYY-MM: {str(e)}'}
                
            # 获取仓库对象
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
                print(f"已找到仓库: ID={warehouse_id}, 名称={warehouse.name}")
            except Warehouse.DoesNotExist:
                print(f"仓库不存在: ID={warehouse_id}")
                return {'error': f'仓库不存在(ID:{warehouse_id})'}
                
            # 检查上月报表是否存在
            try:
                from apps.warehouse.models import Report
                prev_report_date = date(prev_year, prev_month_num, 1)
                prev_report_exists = Report.objects.filter(
                    warehouse=warehouse, 
                    report_date=prev_report_date
                ).exists()
                
                if not prev_report_exists:
                    print(f"上月报表不存在: {prev_month}")
                    # 尝试使用库存数据创建上月报表
                    prev_month_inventory = self._generate_prev_month_inventory(warehouse_id, prev_year, prev_month_num)
                    if prev_month_inventory and len(prev_month_inventory) > 0:
                        print(f"已根据库存数据生成上月报表，物品数量: {len(prev_month_inventory)}")
                        prev_month_data = {
                            'inbound': [],
                            'outbound': [],
                            'inventory': prev_month_inventory
                        }
                        self._save_monthly_report_data(warehouse, prev_year, prev_month_num, prev_month_data)
                    else:
                        print(f"无法生成上月库存数据或物品数量为0")
                else:
                    print(f"已找到上月报表: {prev_month}")
            except Exception as e:
                print(f"检查上月报表出错: {str(e)}")
                import traceback
                print(traceback.format_exc())
            
            # 获取上个月的报表数据
            prev_month_data = self.get_monthly_report_data(warehouse, prev_year, prev_month_num)
            if not prev_month_data:
                print(f"上月报表数据为空: {prev_month}")
                return {'error': f'上个月({prev_month})没有任何报表数据，无法更新本月期初库存'}
                
            if 'inventory' not in prev_month_data:
                print(f"上月报表缺少库存数据: {prev_month}, 报表结构: {list(prev_month_data.keys() if isinstance(prev_month_data, dict) else [])}")
                return {'error': f'上个月({prev_month})没有库存数据，无法更新本月期初库存'}
                
            if not prev_month_data['inventory']:
                print(f"上月报表库存数据为空数组: {prev_month}")
                return {'error': f'上个月({prev_month})库存数据为空，无法更新本月期初库存'}
                
            print(f"成功获取上月库存数据: {prev_month}, 物品数量: {len(prev_month_data['inventory'])}")
                
            # 获取当前月的报表数据
            current_month_data = self.get_monthly_report_data(warehouse, year, month_num)
            if not current_month_data:
                # 如果当前月没有报表数据，则创建一个空的报表数据结构
                print(f"当前月没有报表数据，创建新的空报表: {year}-{month_num}")
                current_month_data = {
                    'inbound': [],
                    'outbound': [],
                    'inventory': []
                }
            else:
                print(f"已找到当前月报表数据: {year}-{month_num}")
                
            # 更新当前月报表的期初库存
            self._update_initial_inventory(current_month_data, prev_month_data)
            
            # 保存更新后的报表数据
            save_result = self._save_monthly_report_data(warehouse, year, month_num, current_month_data)
            if not save_result:
                return {'error': f'保存更新后的报表数据失败'}
                
            return {'success': True, 'message': f'成功更新{year}年{month_num}月的期初库存数据'}
            
        except Exception as e:
            import traceback
            print(f"更新期初库存出错: {str(e)}")
            print(traceback.format_exc())
            return {'error': f'更新期初库存数据出错: {str(e)}'}

    def get_monthly_report_data(self, warehouse, year, month):
        """
        获取指定月份的报表数据
        """
        try:
            from apps.warehouse.models import Report
            
            # 查询月度报表
            report_date = date(year, month, 1)
            try:
                report = Report.objects.get(warehouse=warehouse, report_date=report_date)
                # 如果报表存在但内容为空，返回空对象
                if not report.data or report.data == '{}':
                    print(f"找到报表但数据为空: {year}-{month}")
                    return {}
                print(f"成功获取报表数据: {year}-{month}")
                return report.data
            except Report.DoesNotExist:
                print(f"报表不存在: {year}-{month}")
                # 如果报表不存在，返回空对象
                return {}
                
        except Exception as e:
            print(f"获取月度报表数据出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return {}

    def _save_monthly_report_data(self, warehouse, year, month, data):
        """
        保存月度报表数据
        """
        try:
            from apps.warehouse.models import Report
            
            # 构建报表日期
            report_date = date(year, month, 1)
            
            # 查找或创建报表
            report, created = Report.objects.get_or_create(
                warehouse=warehouse,
                report_date=report_date,
                defaults={
                    'title': f"{warehouse.name}_{year}年{month}月报表",
                    'data': data
                }
            )
            
            # 如果报表已存在，更新内容
            if not created:
                report.title = f"{warehouse.name}_{year}年{month}月报表"
                report.data = data
                report.save()
                print(f"更新已存在的报表: {year}-{month}")
            else:
                print(f"创建新报表: {year}-{month}")
                
            return True
                
        except Exception as e:
            print(f"保存月度报表数据出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return False

    def _update_initial_inventory(self, current_month_data, prev_month_data):
        """
        使用上月期末库存更新当月期初库存
        """
        try:
            # 如果当前月没有库存数据，则初始化
            if 'inventory' not in current_month_data:
                current_month_data['inventory'] = []
                
            print(f"开始更新期初库存: 当前月有 {len(current_month_data['inventory'])} 条记录，上月有 {len(prev_month_data['inventory'])} 条记录")
                
            # 创建当前库存记录的映射，用于快速查找
            current_inventory_map = {}
            for item in current_month_data['inventory']:
                key = f"{item.get('品项')}_{item.get('规格/型号')}"
                current_inventory_map[key] = item
                
            # 处理上月库存数据
            items_updated = 0
            items_added = 0
            
            for prev_item in prev_month_data['inventory']:
                try:
                    product_name = prev_item.get('品项')
                    spec = prev_item.get('规格/型号', '')
                    key = f"{product_name}_{spec}"
                    
                    # 上月期末库存
                    prev_ending_stock = float(prev_item.get('库存', 0))
                    
                    if key in current_inventory_map:
                        # 更新已存在的库存记录
                        current_item = current_inventory_map[key]
                        
                        # 保存原始的累计入库和累计出库数量
                        total_in = float(current_item.get('累计入库', 0))
                        total_out = float(current_item.get('累计出库', 0))
                        
                        # 更新期初库存
                        current_item['期初库存'] = prev_ending_stock
                        
                        # 重新计算当前库存数量
                        current_item['库存'] = prev_ending_stock + total_in - total_out
                        
                        # 更新库存金额
                        current_item['库存金额'] = float(current_item.get('库存', 0)) * float(current_item.get('单价', 0))
                        
                        items_updated += 1
                    else:
                        # 创建新的库存记录
                        new_item = prev_item.copy()
                        new_item['期初库存'] = prev_ending_stock
                        new_item['累计入库'] = 0
                        new_item['累计出库'] = 0
                        new_item['库存'] = prev_ending_stock  # 初始库存等于期初库存
                        
                        current_month_data['inventory'].append(new_item)
                        current_inventory_map[key] = new_item
                        
                        items_added += 1
                except Exception as e:
                    print(f"处理物品 [{product_name}] 的期初库存时出错: {str(e)}")
                    continue
            
            print(f"期初库存更新完成: 更新了 {items_updated} 条记录，新增了 {items_added} 条记录")
            
        except Exception as e:
            print(f"更新期初库存过程中出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            
    def _generate_prev_month_inventory(self, warehouse_id, year, month):
        """
        当上月报表不存在时，尝试根据系统当前库存生成上月库存数据
        """
        try:
            print(f"正在尝试为 {year}-{month} 月生成库存数据...")
            from apps.inventory.models import Inventory
            
            # 获取当前系统中的库存数据
            inventory_records = Inventory.objects.filter(
                warehouse_id=warehouse_id,
                is_active=True
            ).select_related('product', 'location')
            
            if not inventory_records.exists():
                print(f"未找到仓库ID={warehouse_id}的任何库存记录")
                return []
            
            print(f"找到 {inventory_records.count()} 条库存记录，开始生成库存数据")
                
            # 转换为报表格式的库存数据
            inventory_list = []
            for record in inventory_records:
                try:
                    if not record.product:
                        print(f"库存记录 ID={record.id} 没有关联的产品信息，跳过")
                        continue
                        
                    location_code = '未分配'
                    if record.location:
                        location_code = record.location.code
                        
                    # 获取产品单位
                    unit_name = '个'
                    if hasattr(record.product, 'unit') and record.product.unit:
                        unit_name = record.product.unit.name
                        
                    # 处理规格信息    
                    spec = record.spec or (record.product.spec if hasattr(record.product, 'spec') else '')
                    
                    inventory_item = {
                        '序号': len(inventory_list) + 1,
                        '位置': location_code,
                        '品项': record.product.name,
                        '规格/型号': spec,
                        '单位': record.unit or unit_name,
                        '期初库存': float(record.quantity),  # 使用当前库存作为上月期初
                        '累计入库': 0,
                        '累计出库': 0,
                        '库存': float(record.quantity),  # 当前库存
                        '单价': float(record.unit_price),
                        '库存金额': float(record.amount)
                    }
                    
                    inventory_list.append(inventory_item)
                    
                except Exception as e:
                    print(f"处理单个库存记录时出错: {str(e)}")
                    continue
                
            print(f"成功生成库存记录 {len(inventory_list)} 条")
            return inventory_list
            
        except Exception as e:
            print(f"生成上月库存数据出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return []

    @action(detail=True, methods=['post'])
    def update_initial_stock(self, request, pk=None):
        """
        API端点：手动更新指定仓库指定月份的期初库存数据
        
        使用上月的期末库存数据更新当月的期初库存
        """
        try:
            warehouse = self.get_object()
            month = request.data.get('month')
            
            if not month:
                return Response({'error': '缺少月份参数'}, status=status.HTTP_400_BAD_REQUEST)
                
            # 调用内部方法更新期初库存
            result = self.update_initial_inventory_for_month(warehouse.id, month)
            
            if 'error' in result:
                return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)
                
            return Response({
                'message': result['message'],
                'success': True
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            print(f"更新期初库存失败: {str(e)}")
            print(traceback.format_exc())
            return Response({'error': f'更新期初库存失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def backup(self, request, pk=None):
        """
        备份仓库数据
        """
        try:
            logger.info(f"开始备份仓库数据处理: warehouse_id={pk}, 请求来源={request.META.get('REMOTE_ADDR')}")
            warehouse = self.get_object()
            logger.info(f"成功获取仓库信息: warehouse_id={pk}, name={warehouse.name}, code={warehouse.code}")
            
            # 获取仓库相关的所有数据
            data = {
                'warehouse': WarehouseSerializer(warehouse).data,
                'inventory': [],
                'transactions': [],
                'reports': [],
                'backup_time': timezone.now().isoformat()
            }
            logger.info(f"已初始化备份数据结构: warehouse_id={pk}")
            
            # 获取库存数据
            try:
                logger.info(f"开始查询库存数据: warehouse_id={pk}")
                inventory_items = Inventory.objects.filter(
                    warehouse=warehouse,
                    is_active=True  # 使用 is_active 字段
                ).select_related('product', 'location')
                
                logger.info(f"查询到库存记录数: warehouse_id={pk}, count={inventory_items.count()}")
                
                data['inventory'] = []
                inventory_count = 0
                error_count = 0
                
                for item in inventory_items:
                    try:
                        # 安全地获取产品信息
                        product = item.product
                        if not product:
                            logger.warning(f"库存记录 {item.id} 没有关联的产品信息, warehouse_id={pk}")
                            continue
                            
                        # 获取产品属性
                        spec = ''
                        unit = '个'
                        try:
                            spec = product.spec
                        except AttributeError:
                            logger.debug(f"产品 {product.id} 没有spec属性，尝试使用specification属性")
                            try:
                                spec = product.specification
                            except AttributeError:
                                logger.debug(f"产品 {product.id} 没有specification属性")
                                pass
                                
                        try:
                            if hasattr(product, 'unit') and isinstance(product.unit, str):
                                unit = product.unit
                            elif hasattr(product, 'unit') and hasattr(product.unit, 'name'):
                                unit = product.unit.name
                        except Exception as e:
                            logger.debug(f"获取产品 {product.id} 单位信息出错: {str(e)}")
                            pass
                            
                        inventory_data = {
                            'location': item.location.code if item.location else None,
                            'product': product.name,
                            'specification': spec,
                            'unit': unit,
                            'quantity': str(item.quantity or 0),
                            'unit_price': str(item.unit_price or 0),
                            'total_value': str(item.amount or 0),  # 使用 amount 字段
                        }
                        data['inventory'].append(inventory_data)
                        inventory_count += 1
                        if inventory_count % 100 == 0:
                            logger.info(f"已处理 {inventory_count} 条库存记录")
                    except Exception as e:
                        error_count += 1
                        logger.error(f"处理单个库存记录时出错: inventory_id={item.id}, error={str(e)}", exc_info=True)
                        continue
                        
                logger.info(f"成功获取库存数据: warehouse_id={pk}, 成功={inventory_count}条, 失败={error_count}条")
            except Exception as e:
                logger.error(f"获取库存数据时出错: warehouse_id={pk}, error={str(e)}", exc_info=True)
                data['inventory'] = []
            
            # 获取交易记录
            try:
                logger.info(f"开始查询交易记录: warehouse_id={pk}")
                transactions = Transaction.objects.filter(
                    warehouse=warehouse,
                    status='completed'  # 使用 status 字段
                ).select_related('product')
                
                logger.info(f"查询到交易记录数: warehouse_id={pk}, count={transactions.count()}")
                
                data['transactions'] = []
                transaction_count = 0
                error_count = 0
                
                for transaction in transactions:
                    try:
                        # 安全地获取产品信息
                        product = transaction.product
                        if not product:
                            logger.warning(f"交易记录 {transaction.id} 没有关联的产品信息")
                            continue
                            
                        # 获取产品属性
                        spec = ''
                        unit = '个'
                        try:
                            spec = product.spec
                        except AttributeError:
                            try:
                                spec = product.specification
                            except AttributeError:
                                pass
                                
                        try:
                            if hasattr(product, 'unit') and isinstance(product.unit, str):
                                unit = product.unit
                            elif hasattr(product, 'unit') and hasattr(product.unit, 'name'):
                                unit = product.unit.name
                        except Exception as e:
                            logger.debug(f"获取产品 {product.id} 单位信息出错: {str(e)}")
                            pass
                            
                        transaction_data = {
                            'date': transaction.transaction_date.isoformat() if transaction.transaction_date else None,
                            'type': transaction.transaction_type,
                            'product': product.name,
                            'specification': spec,
                            'unit': unit,
                            'quantity': str(transaction.quantity or 0),
                            'unit_price': str(transaction.unit_price or 0),
                            'total_value': str(transaction.amount or 0),  # 使用 amount 字段
                            'handler': transaction.operator.username if transaction.operator else '',  # 使用 operator 字段
                        }
                        data['transactions'].append(transaction_data)
                        transaction_count += 1
                        if transaction_count % 100 == 0:
                            logger.info(f"已处理 {transaction_count} 条交易记录")
                    except Exception as e:
                        error_count += 1
                        logger.error(f"处理单个交易记录时出错: transaction_id={transaction.id}, error={str(e)}", exc_info=True)
                        continue
                        
                logger.info(f"成功获取交易记录: warehouse_id={pk}, 成功={transaction_count}条, 失败={error_count}条")
            except Exception as e:
                logger.error(f"获取交易记录时出错: warehouse_id={pk}, error={str(e)}", exc_info=True)
                data['transactions'] = []
            
            # 获取报表数据
            try:
                logger.info(f"开始获取报表数据: warehouse_id={pk}")
                reports = Report.objects.filter(warehouse=warehouse)
                data['reports'] = ReportSerializer(reports, many=True).data
                logger.info(f"成功获取报表数据: warehouse_id={pk}, 共{len(data['reports'])}条记录")
            except Exception as e:
                logger.error(f"获取报表数据时出错: warehouse_id={pk}, error={str(e)}", exc_info=True)
                data['reports'] = []
            
            # 将数据转换为JSON并返回
            try:
                logger.info(f"开始生成JSON备份文件: warehouse_id={pk}")
                try:
                    data_size = len(str(data))
                    logger.info(f"备份数据大小估计: warehouse_id={pk}, size={data_size}字节")
                except Exception as size_e:
                    logger.warning(f"无法估计备份数据大小: {str(size_e)}")
                
                json_data = json.dumps(data, ensure_ascii=False, indent=2)
                logger.info(f"JSON序列化完成: warehouse_id={pk}, size={len(json_data)}字节")
                
                response = HttpResponse(json_data, content_type='application/json')
                filename = f'warehouse_{warehouse.code}_backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                # 添加额外的响应头，确保浏览器正确处理
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                
                logger.info(f"备份数据生成成功: warehouse_id={pk}, filename={filename}, size={len(json_data)}字节")
                return response
            except Exception as e:
                logger.error(f"生成备份文件时出错: warehouse_id={pk}, error={str(e)}", exc_info=True)
                return Response(
                    {'error': f'生成备份文件失败: {str(e)}'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
        except Http404:
            logger.error(f"仓库不存在: warehouse_id={pk}")
            return Response(
                {'error': '仓库不存在'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"备份仓库数据失败: warehouse_id={pk}, error={str(e)}", exc_info=True)
            import traceback
            trace = traceback.format_exc()
            logger.error(f"详细错误堆栈: {trace}")
            return Response(
                {'error': f'备份失败: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """
        从备份恢复仓库数据
        """
        try:
            warehouse = self.get_object()
            backup_data = request.data
            
            if not isinstance(backup_data, dict):
                return Response({'error': '无效的备份数据格式'}, status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                # 验证备份数据中的仓库信息
                if backup_data.get('warehouse', {}).get('code') != warehouse.code:
                    return Response({'error': '备份数据与目标仓库不匹配'}, status=status.HTTP_400_BAD_REQUEST)
                
                # 清除现有数据
                Inventory.objects.filter(warehouse=warehouse).delete()
                Transaction.objects.filter(warehouse=warehouse).delete()
                Report.objects.filter(warehouse=warehouse).delete()
                
                # 恢复库存数据
                for item in backup_data.get('inventory', []):
                    location = None
                    if item.get('location'):
                        location = WarehouseLocation.objects.filter(code=item['location'], warehouse=warehouse).first()
                    
                    product = Product.objects.filter(name=item['product']).first()
                    if not product:
                        product = Product.objects.create(
                            name=item['product'],
                            specification=item.get('specification', ''),
                            unit=item.get('unit', '个')
                        )
                    
                    Inventory.objects.create(
                        warehouse=warehouse,
                        location=location,
                        product=product,
                        quantity=item['quantity'],
                        unit_price=item['unit_price'],
                        total_value=item['total_value']
                    )
                
                # 恢复交易记录
                for trans in backup_data.get('transactions', []):
                    product = Product.objects.filter(name=trans['product']).first()
                    if not product:
                        product = Product.objects.create(
                            name=trans['product'],
                            specification=trans.get('specification', ''),
                            unit=trans.get('unit', '个')
                        )
                    
                    Transaction.objects.create(
                        warehouse=warehouse,
                        transaction_date=timezone.parse_datetime(trans['date']) if trans.get('date') else None,
                        transaction_type=trans['type'],
                        product=product,
                        quantity=trans['quantity'],
                        unit_price=trans['unit_price'],
                        total_value=trans['total_value'],
                        handler=trans['handler']
                    )
                
                # 恢复报表数据
                for report_data in backup_data.get('reports', []):
                    report_data['warehouse'] = warehouse.id
                    serializer = ReportSerializer(data=report_data)
                    if serializer.is_valid():
                        serializer.save()
                
                return Response({'message': '恢复备份成功'}, status=status.HTTP_200_OK)
                
        except Exception as e:
            logger.error(f"恢复备份失败: {str(e)}")
            return Response({'error': f'恢复失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class WarehouseAreaViewSet(viewsets.ModelViewSet):
    queryset = WarehouseArea.objects.all()
    serializer_class = WarehouseAreaSerializer
    permission_classes = [WarehouseViewPermission]
    filterset_fields = ['code', 'name', 'warehouse', 'area_type']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['id', 'code', 'name', 'created_time']
    ordering = ['-created_time']

class WarehouseLocationViewSet(viewsets.ModelViewSet):
    queryset = WarehouseLocation.objects.all()
    serializer_class = WarehouseLocationSerializer
    permission_classes = [WarehouseViewPermission]
    filterset_fields = ['code', 'name', 'area', 'location_type', 'is_empty']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['id', 'code', 'name', 'created_time']
    ordering = ['-created_time']

class ReportViewSet(viewsets.ModelViewSet):
    """
    报表管理视图集
    提供报表的CRUD功能
    """
    queryset = Report.objects.all().order_by('-created_at')
    serializer_class = ReportSerializer
    permission_classes = [WarehouseViewPermission]
    filterset_fields = ['warehouse', 'report_date', 'creator']
    search_fields = ['title', 'description']

    def get_queryset(self):
        """根据查询参数过滤报表"""
        queryset = super().get_queryset()
        
        # 按仓库过滤
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
            
        # 按月份过滤
        month = self.request.query_params.get('month')
        if month:
            try:
                year, month = month.split('-')
                queryset = queryset.filter(
                    report_date__year=int(year), 
                    report_date__month=int(month)
                )
            except ValueError:
                pass
                
        return queryset

    def perform_create(self, serializer):
        """创建报表时设置创建者"""
        serializer.save(creator=self.request.user)

class MonthlyReportViewSet(viewsets.ViewSet):
    """
    月度报表视图集
    提供月度报表的各种操作功能
    """
    permission_classes = [WarehouseViewPermission]
    
    # 添加基本查询方法
    def list(self, request):
        """获取月度报表列表"""
        try:
            warehouse_id = request.query_params.get('warehouse_id')
            month = request.query_params.get('month')
            
            if not warehouse_id or not month:
                return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                year, month_num = month.split('-')
                year = int(year)
                month_num = int(month_num)
            except (ValueError, IndexError):
                return Response({'error': '月份格式错误，应为YYYY-MM'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({'error': '仓库不存在'}, status=status.HTTP_404_NOT_FOUND)
            
            # 获取月度报表数据
            monthly_data = self.get_monthly_report_data(warehouse, year, month_num)
            
            # 如果没有数据，返回空对象而不是错误
            if not monthly_data:
                monthly_data = {
                    'inbound': [],
                    'outbound': [],
                    'inventory': []
                }
            
            return Response(monthly_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            print(f"获取月度报表失败: {str(e)}")
            print(traceback.format_exc())
            return Response({'error': f'获取月度报表失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def update_initial_stock(self, request):
        """
        API端点：手动更新指定月份的期初库存数据
        
        使用上月的期末库存数据更新当月的期初库存
        """
        try:
            warehouse_id = request.data.get('warehouse_id')
            month = request.data.get('month')
            
            if not warehouse_id or not month:
                return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)
                
            # 调用内部方法更新期初库存
            result = self.update_initial_inventory_for_month(warehouse_id, month)
            
            if 'error' in result:
                return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)
                
            return Response({
                'message': result['message'],
                'success': True
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            print(f"更新期初库存失败: {str(e)}")
            print(traceback.format_exc())
            return Response({'error': f'更新期初库存失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_monthly_report_data(self, warehouse, year, month):
        """
        获取指定月份的报表数据
        """
        try:
            from apps.warehouse.models import Report
            
            # 查询月度报表
            report_date = date(year, month, 1)
            try:
                report = Report.objects.get(warehouse=warehouse, report_date=report_date)
                # 如果报表存在但内容为空，返回空对象
                if not report.data or report.data == '{}':
                    print(f"找到报表但数据为空: {year}-{month}")
                    return {}
                print(f"成功获取报表数据: {year}-{month}")
                return report.data
            except Report.DoesNotExist:
                print(f"报表不存在: {year}-{month}")
                # 如果报表不存在，返回空对象
                return {}
                
        except Exception as e:
            print(f"获取月度报表数据出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return {}

    def _save_monthly_report_data(self, warehouse, year, month, data):
        """
        保存月度报表数据
        """
        try:
            from apps.warehouse.models import Report
            
            # 构建报表日期
            report_date = date(year, month, 1)
            
            # 查找或创建报表
            report, created = Report.objects.get_or_create(
                warehouse=warehouse,
                report_date=report_date,
                defaults={
                    'title': f"{warehouse.name}_{year}年{month}月报表",
                    'data': data
                }
            )
            
            # 如果报表已存在，更新内容
            if not created:
                report.title = f"{warehouse.name}_{year}年{month}月报表"
                report.data = data
                report.save()
                print(f"更新已存在的报表: {year}-{month}")
            else:
                print(f"创建新报表: {year}-{month}")
                
            return True
                
        except Exception as e:
            print(f"保存月度报表数据出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return False

    def _update_initial_inventory(self, current_month_data, prev_month_data):
        """
        使用上月期末库存更新当月期初库存
        """
        try:
            # 如果当前月没有库存数据，则初始化
            if 'inventory' not in current_month_data:
                current_month_data['inventory'] = []
                
            print(f"开始更新期初库存: 当前月有 {len(current_month_data['inventory'])} 条记录，上月有 {len(prev_month_data['inventory'])} 条记录")
                
            # 创建当前库存记录的映射，用于快速查找
            current_inventory_map = {}
            for item in current_month_data['inventory']:
                key = f"{item.get('品项')}_{item.get('规格/型号')}"
                current_inventory_map[key] = item
                
            # 处理上月库存数据
            items_updated = 0
            items_added = 0
            
            for prev_item in prev_month_data['inventory']:
                try:
                    product_name = prev_item.get('品项')
                    spec = prev_item.get('规格/型号', '')
                    key = f"{product_name}_{spec}"
                    
                    # 上月期末库存
                    prev_ending_stock = float(prev_item.get('库存', 0))
                    
                    if key in current_inventory_map:
                        # 更新已存在的库存记录
                        current_item = current_inventory_map[key]
                        
                        # 保存原始的累计入库和累计出库数量
                        total_in = float(current_item.get('累计入库', 0))
                        total_out = float(current_item.get('累计出库', 0))
                        
                        # 更新期初库存
                        current_item['期初库存'] = prev_ending_stock
                        
                        # 重新计算当前库存数量
                        current_item['库存'] = prev_ending_stock + total_in - total_out
                        
                        # 更新库存金额
                        current_item['库存金额'] = float(current_item.get('库存', 0)) * float(current_item.get('单价', 0))
                        
                        items_updated += 1
                    else:
                        # 创建新的库存记录
                        new_item = prev_item.copy()
                        new_item['期初库存'] = prev_ending_stock
                        new_item['累计入库'] = 0
                        new_item['累计出库'] = 0
                        new_item['库存'] = prev_ending_stock  # 初始库存等于期初库存
                        
                        current_month_data['inventory'].append(new_item)
                        current_inventory_map[key] = new_item
                        
                        items_added += 1
                except Exception as e:
                    print(f"处理物品 [{product_name}] 的期初库存时出错: {str(e)}")
                    continue
            
            print(f"期初库存更新完成: 更新了 {items_updated} 条记录，新增了 {items_added} 条记录")
            
        except Exception as e:
            print(f"更新期初库存过程中出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            
    def _generate_prev_month_inventory(self, warehouse_id, year, month):
        """
        当上月报表不存在时，尝试根据系统当前库存生成上月库存数据
        """
        try:
            print(f"正在尝试为 {year}-{month} 月生成库存数据...")
            from apps.inventory.models import Inventory
            
            # 获取当前系统中的库存数据
            inventory_records = Inventory.objects.filter(
                warehouse_id=warehouse_id,
                is_active=True
            ).select_related('product', 'location')
            
            if not inventory_records.exists():
                print(f"未找到仓库ID={warehouse_id}的任何库存记录")
                return []
            
            print(f"找到 {inventory_records.count()} 条库存记录，开始生成库存数据")
                
            # 转换为报表格式的库存数据
            inventory_list = []
            for record in inventory_records:
                try:
                    if not record.product:
                        print(f"库存记录 ID={record.id} 没有关联的产品信息，跳过")
                        continue
                        
                    location_code = '未分配'
                    if record.location:
                        location_code = record.location.code
                        
                    # 获取产品单位
                    unit_name = '个'
                    if hasattr(record.product, 'unit') and record.product.unit:
                        unit_name = record.product.unit.name
                        
                    # 处理规格信息    
                    spec = record.spec or (record.product.spec if hasattr(record.product, 'spec') else '')
                    
                    inventory_item = {
                        '序号': len(inventory_list) + 1,
                        '位置': location_code,
                        '品项': record.product.name,
                        '规格/型号': spec,
                        '单位': record.unit or unit_name,
                        '期初库存': float(record.quantity),  # 使用当前库存作为上月期初
                        '累计入库': 0,
                        '累计出库': 0,
                        '库存': float(record.quantity),  # 当前库存
                        '单价': float(record.unit_price),
                        '库存金额': float(record.amount)
                    }
                    
                    inventory_list.append(inventory_item)
                    
                except Exception as e:
                    print(f"处理单个库存记录时出错: {str(e)}")
                    continue
                
            print(f"成功生成库存记录 {len(inventory_list)} 条")
            return inventory_list
            
        except Exception as e:
            print(f"生成上月库存数据出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return []

    def update_initial_inventory_for_month(self, warehouse_id, month):
        """
        更新指定月份报表的期初库存数据，基于上月的期末库存
        
        参数:
            warehouse_id: 仓库ID
            month: 目标月份，格式：YYYY-MM
        """
        try:
            # 参数校验
            if not warehouse_id or not month:
                return {'error': '缺少必要参数'}
                
            # 解析月份
            try:
                year, month_num = month.split('-')
                year = int(year)
                month_num = int(month_num)
                
                # 计算上个月的日期
                first_day_of_month = datetime(year, month_num, 1)
                last_day_of_prev_month = first_day_of_month - timedelta(days=1)
                prev_month = last_day_of_prev_month.strftime("%Y-%m")
                prev_year = last_day_of_prev_month.year
                prev_month_num = last_day_of_prev_month.month
                
                print(f"正在处理月份: 当前月={year}-{month_num}, 上个月={prev_year}-{prev_month_num}")
            except (ValueError, IndexError) as e:
                print(f"月份格式解析错误: {str(e)}")
                return {'error': f'月份格式错误，应为YYYY-MM: {str(e)}'}
                
            # 获取仓库对象
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
                print(f"已找到仓库: ID={warehouse_id}, 名称={warehouse.name}")
            except Warehouse.DoesNotExist:
                print(f"仓库不存在: ID={warehouse_id}")
                return {'error': f'仓库不存在(ID:{warehouse_id})'}
                
            # 检查上月报表是否存在
            try:
                from apps.warehouse.models import Report
                prev_report_date = date(prev_year, prev_month_num, 1)
                prev_report_exists = Report.objects.filter(
                    warehouse=warehouse, 
                    report_date=prev_report_date
                ).exists()
                
                if not prev_report_exists:
                    print(f"上月报表不存在: {prev_month}")
                    # 尝试使用库存数据创建上月报表
                    prev_month_inventory = self._generate_prev_month_inventory(warehouse_id, prev_year, prev_month_num)
                    if prev_month_inventory and len(prev_month_inventory) > 0:
                        print(f"已根据库存数据生成上月报表，物品数量: {len(prev_month_inventory)}")
                        prev_month_data = {
                            'inbound': [],
                            'outbound': [],
                            'inventory': prev_month_inventory
                        }
                        self._save_monthly_report_data(warehouse, prev_year, prev_month_num, prev_month_data)
                    else:
                        print(f"无法生成上月库存数据或物品数量为0")
                else:
                    print(f"已找到上月报表: {prev_month}")
            except Exception as e:
                print(f"检查上月报表出错: {str(e)}")
                import traceback
                print(traceback.format_exc())
            
            # 获取上个月的报表数据
            prev_month_data = self.get_monthly_report_data(warehouse, prev_year, prev_month_num)
            if not prev_month_data:
                print(f"上月报表数据为空: {prev_month}")
                return {'error': f'上个月({prev_month})没有任何报表数据，无法更新本月期初库存'}
                
            if 'inventory' not in prev_month_data:
                print(f"上月报表缺少库存数据: {prev_month}, 报表结构: {list(prev_month_data.keys() if isinstance(prev_month_data, dict) else [])}")
                return {'error': f'上个月({prev_month})没有库存数据，无法更新本月期初库存'}
                
            if not prev_month_data['inventory']:
                print(f"上月报表库存数据为空数组: {prev_month}")
                return {'error': f'上个月({prev_month})库存数据为空，无法更新本月期初库存'}
                
            print(f"成功获取上月库存数据: {prev_month}, 物品数量: {len(prev_month_data['inventory'])}")
                
            # 获取当前月的报表数据
            current_month_data = self.get_monthly_report_data(warehouse, year, month_num)
            if not current_month_data:
                # 如果当前月没有报表数据，则创建一个空的报表数据结构
                print(f"当前月没有报表数据，创建新的空报表: {year}-{month_num}")
                current_month_data = {
                    'inbound': [],
                    'outbound': [],
                    'inventory': []
                }
            else:
                print(f"已找到当前月报表数据: {year}-{month_num}")
                
            # 更新当前月报表的期初库存
            self._update_initial_inventory(current_month_data, prev_month_data)
            
            # 保存更新后的报表数据
            save_result = self._save_monthly_report_data(warehouse, year, month_num, current_month_data)
            if not save_result:
                return {'error': f'保存更新后的报表数据失败'}
                
            return {'success': True, 'message': f'成功更新{year}年{month_num}月的期初库存数据'}
            
        except Exception as e:
            import traceback
            print(f"更新期初库存出错: {str(e)}")
            print(traceback.format_exc())
            return {'error': f'更新期初库存数据出错: {str(e)}'}